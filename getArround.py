# -*- coding: utf-8 -*-
#依据酒店id获取位置数据
import requests
import json
import openpyxl
import logging
from concurrent.futures import ThreadPoolExecutor
import random
import time
from openpyxl import load_workbook

filename1 = 'G:\PythonFIle\meituan\code\data\hotelList.xlsx'
filename2 = 'G:\PythonFIle\meituan\code\data\hotelA6.xlsx'
cate= [6,3,0,1,4] #虽然有5个类型，但有一个类型不可用，因此获取了车站、购物、餐饮、景点数据
poiCate = 6


ids = []
cid = []

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
url = 'https://ihotel.meituan.com/detailapi/api/around/info'
wb2 = load_workbook(filename1)
ws = wb2['Sheet1']
row_max = ws.max_row
con_max = ws.max_column
for j in ws.rows: # we.rows 获取每一行数据
    n = j[0]
    c = j[13]
    ids.append(n.value)
    cid.append(c.value)
    print(n.value,c.value)
print(len(ids),len(cid))
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['hotel_id','poi_id','directUrl', 'distance', 'distanceDouble', 'landMarkId', 'latitude', 'longitude', 'poiName', 'posDec', 'score', 'type'])

def hotel_data(x):
    #print(ids[x],cid[x])
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'Cookie': '_lxsdk_cuid=177677169faa3-0b79cb49190f37-7711a3e-144000-177677169fbc8; ci=57; rvct=57%2C52; iuuid=148E5E533C30EE4E474C84F1D580BA182639D74FB94A1C4B8658499422FD2D72; cityname=%E6%AD%A6%E6%B1%89; _lxsdk=148E5E533C30EE4E474C84F1D580BA182639D74FB94A1C4B8658499422FD2D72; uuid=e2fec2b2b4504afe91c5.1613653923.1.0.0; mtcdn=K; lsu=; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; IJSESSIONID=1t5gvq1c85t9h1vsmw1hsutbxb; _lxsdk_s=177b9d093d6-e09-d90-b9b%7C%7C18',
        'Host': 'ihotel.meituan.com',
        'Origin': 'https://hotel.meituan.com',
        'Referer': 'https://hotel.meituan.com/'+str(ids[x])+'/?ci=2021-03-04&co=2021-03-04',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
    }
    data = {
        'poiCategory': poiCate,
        'poiid': ids[x],
        'cityId': cid[x],
        'X-FOR-WITH': '5doLwOwvyaOWNZBMZR%2BptaF2sG98NUTU8yWmGmh1O1qcsO3H8rW50c0pJ%2B5L4CpAaAIuk2XnaoD60zgvKDZ%2F%2FY2fdndG5qewicPNoJlgvdimz%2FaElmi4ur5MVTYdJuyFMhSab3z8QW80MCJ1dnGqowg3yW4clT6DQQox02qbPEuMKmVE71B0jNqr5aRVUaOgr0mrP1zUpaVAhQk9dMjn0AiMRLl5yWkI2nC38rvi86c%3D',
    }
    res = requests.get(url, headers=headers, params=data)
    #print(res.url)
    time.sleep(random.randint(1, 3))
    results = json.loads(res.text)['data']['aroundPoiInfoModelList']
    for con in results:
        directUrl = con['directUrl']
        distance = con['distance']
        distanceDouble = con['distanceDouble']
        landMarkId = con['landMarkId']
        latitude = con['latitude']
        longitude = con['longitude']
        poiName = con['poiName']
        posDec = con['posDec']
        score = con['score']
        type = con['type']
        poiId= con['poiId']
        hotel_id = ids[x]
        data = [hotel_id,poiId,directUrl, distance, distanceDouble, landMarkId, latitude, longitude, poiName, posDec, score, type]
        sheet.append(data)
        logging.info(data)
        wb.save(filename2)

if __name__ == '__main__':
    order = [i for i in range(len(ids))]
    print(range(len(ids)))
    with ThreadPoolExecutor( max_workers=16) as executor:
        executor.map(hotel_data, order)
