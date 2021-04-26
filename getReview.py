# -*- coding: utf-8 -*-
#依据酒店id获取评论数据
import pandas as pd
import urllib.request as req
import json
import sys
import time
import random
import re
from openpyxl import load_workbook

ids = []
wb2 = load_workbook('/code/data/hotelList.xlsx')
ws = wb2['Sheet']
row_max = ws.max_row
con_max = ws.max_column
for j in ws.rows:
    n=j[0]
    print(n.value)
    ids.append(n.value)
print(len(ids))
print(sys.getdefaultencoding())

class MTCommentsCrawler:
    def __init__(self, productId=None, limit=10, start=0):
        self.productId = productId  # 酒店ID
        self.limit = limit  # 一次获取多少条评论
        self.start = start
        self.locationLink = 'https://ihotel.meituan.com/api/v2/comments/biz/reviewList'
        self.paramValue = {
            'referid': self.productId,
            'limit': self.limit,
            'start': self.start,
        }
        self.locationUrl = None

    # 构造url调用参数
    def paramDict2Str(self, params):
        str1 = ''
        for p, v in params.items():
            str1 = str1 + p + '=' + str(v) + '&'
        return str1

    # 构造调用url
    def concatLinkParam(self):
        self.locationUrl = self.locationLink + '?' + self.paramDict2Str(
            self.paramValue) + 'filterid=800&querytype=1&utm_medium=touch&version_name=999.9'
        # print(self.locationUrl)

    # 伪装浏览器进行数据请求
    def requestMethodPage(self):
        # 伪装浏览器 ，打开网站
        headers = {
            'Connection': 'Keep-Alive',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Mobile Safari/537.36',
            'Referer': 'https://i.meituan.com/awp/h5/hotel-v2/feedback/index.html?poiId=%d' % (self.productId),
            'Host': 'ihotel.meituan.com'
        }
        url = self.locationUrl
        print('url : ', url)
        reqs = req.Request(url, headers=headers)
        return reqs

    # 读取服务端获取的数据，返回json格式
    def showListPage(self):
        request_m = self.requestMethodPage()
        conn = req.urlopen(request_m)
        return_str = conn.read().decode('utf-8')
        return json.loads(return_str)

    # 将评论数据保存到本地
    def save_csv(self, df):
        # 保存文件
        df.to_csv(path_or_buf='mt_%d.csv' % self.productId, sep=',', header=True, index=True,
                  encoding='utf_8_sig')

    # 移除换行符，#，表情
    def remove_emoji(self, text):
        text = text.replace('\n', '')
        text = text.replace('#', '')
        try:
            highpoints = re.compile(u'[\U00010000-\U0010ffff]')
        except re.error:
            highpoints = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
        return highpoints.sub(u'', text)

    # 抓取数据
    def crawler(self):
        # 把抓取的数据存入CSV文件，设置时间间隔，以免被屏蔽
        json_info = self.showListPage()
        tmp_list = []
        tmp_text_list = []
        # print(json_info)
        Data = json_info['Data']
        comments = Data['List']
        for com in comments:
            text = self.remove_emoji(com['Content'])
            star = com['Star']
            tmp_list.append([star, text])
            tmp_text_list.append([text])
        df = pd.DataFrame(tmp_list, columns=['star', 'content'])
        self.save_csv(df)  # 保存为csv


# 初始化参数
def mtComment():
    productIdGroup = ids  # 酒店ID组
    limit = 60
    for productId in productIdGroup:
        start = random.randint(1, 9)
        MTC = MTCommentsCrawler(productId, limit, start)
        MTC.concatLinkParam()
        MTC.crawler()
        time.sleep(random.randint(1, 5))


if __name__ == '__main__':
    mtComment()
